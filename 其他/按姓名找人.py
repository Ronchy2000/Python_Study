# -*- coding: utf-8 -*-
# @Time    : 2021/10/11 19:18
# @Author  : Ronchy_LU
# @Email   : rongqi1949@gmail.com
# @File    : 按姓名找人.py
# @Software: PyCharm


import os
import openpyxl
Befound_file = "C:\/Users\Ronchy\OneDrive\桌面\传感器期中论文\提交名单.xlsx"
BefoundWorkbook = openpyxl.load_workbook(Befound_file)
#获取工作表名
BeFoundWorkName=BefoundWorkbook.sheetnames
print("工作表名称:",BeFoundWorkName)
#获取工作表对象BeFoundWorksheet
BeFoundWorksheet = BefoundWorkbook[BeFoundWorkName[0]]
print("工作表对象:",BeFoundWorksheet)
#获取工作表的属性
BeFoundName_sheet=BeFoundWorksheet.title
print("sheet:",BeFoundName_sheet)
rows = BeFoundWorksheet.max_row
columns = BeFoundWorksheet.max_column
print("rows:",rows,"columns:",columns)

'''访问列数据时容易出错：
        sheet.rows/columns是生成器类型，不能使用索引。
        所以我们将其转换为list之后再使用索引
'''
beFound_arry = []
for col in list(BeFoundWorksheet.columns)[0]:
    beFound_arry.append(col.value)
    #print(col.value,end=" ")
    #print(beFound_arry[-1])
#print(type(beFound_arry[0]))

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#

excel_path = "G:\物电七班班级事务\大三工作/2019级电子信息工程7班大三.xlsx"
workbook = openpyxl.load_workbook(excel_path)

#获取工作表名
workname=workbook.sheetnames
print("工作表名称:",workname)
#获取工作表对象sheet1
worksheet1 = workbook[workname[0]]
print("工作表对象:",worksheet1)
#获取工作表的属性
name_sheet=worksheet1.title
print("sheet:",name_sheet)

rows = worksheet1.max_row
columns = worksheet1.max_column
# print(rows,columns)  #57行7列

#按行或列获取表中的数据
'''要想以行方式或者列方式，获取整个工作表的内容，我们需要使用到以下两个生成器：
    sheet.rows，这是一个生成器，里面是每一行的数据，每一行数据由一个元组类型包裹。
    sheet.columns，同上，里面是每一列的数据'''
#按行打印
# for row in worksheet1.rows:
#     for cell in row:
#         print(cell.value,end='')
#     print()

# 按列打印
# for col in worksheet1.columns:
#     for cell in col:
#         print(cell.value,end=" ")
#     print()

#sheet.rows/colums是生成器类型，不能使用索引。所以我们将其转换为list之后再使用索引
#例如用list(sheet.rows)[3]来获取第四行的tuple对象
origin_name = []
for cell in list(worksheet1.columns)[0]:  #获取第1列的数据
    origin_name.append(cell.value)
    #print(origin_name[-1]) #Standard 名单
#print(type(origin_name[0])) #str类型
##解决bug，list中学号，最后None ，把str类型变为int ，或者把上面被查找的list变成str
origin_name.remove("学号")
origin_name.pop(-1)
#print(origin_name)
origin_name = [int(i) for i in origin_name]
#print(type(origin_name[0]))

print("People:",len(origin_name[1:-1])) #55人

for somebody in origin_name[1:-1]:
    if somebody not in beFound_arry:
        print("未找到:",somebody)
# print("End!")