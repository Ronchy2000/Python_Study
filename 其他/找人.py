# -*- coding: utf-8 -*-
# @Time    : 2021/5/23 20:03
# @Author  : Ronchy_LU
# @Email   : rongqi1949@gmail.com
# @File    : 找人.py
# @Software: PyCharm
import os
import openpyxl

file_dir = "F:/1907实验六"
lst = []
llst = []
for files in os.walk(file_dir, topdown=False):
    # print(root)     # 当前目录路径
    # print(dirs)     # 当前目录下所有子目录
    # print(files)    # 当前路径下所有非目录子文件
    lst = files
# print(lst[-1])
for temp in lst[-1]:
    llst.append(temp)
# print(llst[-1])
name = []
#llst是单独元素的列表
for i in llst:
    if i[-1]=="c":  #doc结尾的
        name.append(i[11:-8])
    else:
        name.append(i[11:-9])  #docx结尾的
#print(llst[0][11:-8])    #从11开始
print("提交了的同学:",name)
#name已经是获取到的名字了
###################################
excel_path = "F:/物电七班班级事务/学生信息/2019级电子信息工程7班.xlsx"
workbook = openpyxl.load_workbook(excel_path)

#获取工作表名
workname=workbook.sheetnames
print("工作表名称:",workname)
#获取工作表对象sheet1
worksheet1 = workbook[workname[0]]
print("工作表对象:",worksheet1)
#获取工作表的属性
name_sheet=worksheet1.title
print("sheet:",name_sheet)

rows = worksheet1.max_row
columns = worksheet1.max_column
# print(rows,columns)  #57行7列

#按行或列获取表中的数据
'''要想以行方式或者列方式，获取整个工作表的内容，我们需要使用到以下两个生成器：
    sheet.rows，这是一个生成器，里面是每一行的数据，每一行数据由一个元组类型包裹。
    sheet.columns，同上，里面是每一列的数据'''
#按行打印
# for row in worksheet1.rows:
#     for cell in row:
#         print(cell.value,end='')
#     print()

# 按列打印
# for col in worksheet1.columns:
#     for cell in col:
#         print(cell.value,end=" ")
#     print()

#sheet.rows/colums是生成器类型，不能使用索引。所以我们将其转换为list之后再使用索引
#例如用list(sheet.rows)[3]来获取第四行的tuple对象
origin_name = []
for cell in list(worksheet1.columns)[1]:  #获取第2列的数据
    origin_name.append(cell.value)
    print(origin_name[-1]) #Standard 名单
print(file_dir)
print("People:",len(origin_name[1:-1])) #55人

for std in origin_name[1:-1]:
    if std not in name:
        print("未找到:",std)
print("End!")




